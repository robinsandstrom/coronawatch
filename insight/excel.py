from collections import OrderedDict
from openpyxl import Workbook
from insight.models import CoronaCase, region_codes, CountryTracker
from datetime import datetime, timedelta
from insight.backend import aggregate_by_dates

def get_excel_file():
    wb = Workbook()
    wb.active = 0
    ws = wb.active
    ws.title = 'COVID-19-geographic-distributi'


    setHeader(ws)
    setData(ws)

    return wb

def setHeader(ws):
    headers = ['DateRep', 'Day', 'Month', 'Year', 'Cases', 'Deaths', 'Countries and territories', 'GeoId', 'In hospital care', 'In intensive care']
    for ix in range(1, len(headers)+1):
        _ = ws.cell(column=ix, row=1, value=headers[ix-1])

def setData(ws):
    cases = CountryTracker.objects.all().order_by('country', '-date')
    ix=2
    for case in cases:
        _ = ws.cell(column=1, row=ix, value=excel_date(case.date))
        _ = ws.cell(column=2, row=ix, value=case.date.day)
        _ = ws.cell(column=3, row=ix, value=case.date.month)
        _ = ws.cell(column=4, row=ix, value=case.date.year)
        _ = ws.cell(column=5, row=ix, value=case.new_cases)
        _ = ws.cell(column=6, row=ix, value=case.new_deaths)
        _ = ws.cell(column=7, row=ix, value=case.country)
        _ = ws.cell(column=8, row=ix, value=case.country)
        _ = ws.cell(column=9, row=ix, value=case.in_hospital)
        _ = ws.cell(column=10, row=ix, value=case.in_intensive_care)
        ix+=1

def parse_for_excel(cases):
    regional_data = {}

    for c in region_codes:
        r_cases = cases.filter(region=c)
        regional_data[c] = aggregate_by_dates(r_cases, days=365)

    return regional_data

def excel_date(date1):
    temp = datetime(1899, 12, 30).date()    # Note, not 31st Dec but 30th!
    delta = date1 - temp
    return float(delta.days) + (float(delta.seconds) / 86400)
